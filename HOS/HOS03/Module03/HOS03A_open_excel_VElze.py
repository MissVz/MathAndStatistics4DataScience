# CS506 - HOS03 - Web Scraping and Spreadsheets
# Script 5: read_excel.py 📊
# Purpose: Read and format data from an Excel spreadsheet

import openpyxl

# Load the Excel workbook
wb = openpyxl.load_workbook('./example.xlsx')
print(wb.sheetnames)

# Access the first sheet
sheet = wb.active
print(sheet.title)

# Read rows from columns A, B, C
for i in range(1, sheet.max_row + 1):
    # ljust is used for clean formatting
    print(
        i,
        str(sheet.cell(row=i, column=1).value).ljust(25),
        str(sheet.cell(row=i, column=2).value.ljust(15, ' ')),
        str(sheet.cell(row=i, column=3).value)
    )

# REFERENCE: OpenAI. (2025). ChatGPT’s assistance with openpyxl spreadsheet reading [Large language model]. https://openai.com/chatgpt
